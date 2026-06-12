from __future__ import annotations

import csv
import io
import re
import unicodedata
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from .cnpj import extract_cnpjs


def _decode(data: bytes) -> str:
    # utf-8 variants first (correct for most exports), then GBK/GB18030 for the
    # Chinese-sourced CSVs, and latin-1 last because it never raises.
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="ignore")


def _clean(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()


def _normalize_header(text: str) -> str:
    decomposed = unicodedata.normalize("NFKD", str(text or ""))
    stripped = "".join(ch for ch in decomposed if not unicodedata.combining(ch))
    return re.sub(r"[\s_\-/.]+", "", stripped.strip().casefold())


# Header aliases (already normalized via _normalize_header) used to locate the
# company-name column and its optional companions when a file has no CNPJs.
_HEADER_ALIASES: dict[str, set[str]] = {
    "name": {
        "公司名称", "公司名", "公司", "companyname", "company", "empresa",
        "razaosocial", "razao", "nome", "name", "companhia", "company名称",
    },
    "website": {"公司网址", "网址", "网站", "website", "url", "site", "homepage"},
    "email": {"邮箱地址", "邮箱", "邮件", "email", "mail", "emailaddress"},
    "responsible": {
        "法人", "法定代表人", "法人代表", "responsavel", "socio", "owner",
        "representante", "responsible",
    },
}


@dataclass
class NameQuery:
    company_name: str
    website: str = ""
    email: str = ""
    responsible_hint: str = ""
    sheet_name: str = ""
    row_number: int = 0

    def to_dict(self) -> dict:
        return {
            "company_name": self.company_name,
            "website": self.website,
            "email": self.email,
            "responsible_hint": self.responsible_hint,
            "sheet_name": self.sheet_name,
            "row_number": self.row_number,
        }

    @classmethod
    def from_dict(cls, data: dict | None) -> "NameQuery":
        payload = data or {}
        return cls(
            company_name=str(payload.get("company_name", "")),
            website=str(payload.get("website", "")),
            email=str(payload.get("email", "")),
            responsible_hint=str(payload.get("responsible_hint", "")),
            sheet_name=str(payload.get("sheet_name", "")),
            row_number=int(payload.get("row_number", 0) or 0),
        )


@dataclass
class UploadDetails:
    filename: str
    cnpjs: list[str]
    row_refs: list[dict] = field(default_factory=list)
    source_type: str = ""
    mode: str = "cnpj"
    name_queries: list[NameQuery] = field(default_factory=list)


def _detect_name_columns(header: list) -> dict[str, int]:
    columns: dict[str, int] = {}
    for index, raw in enumerate(header):
        key = _normalize_header(raw)
        if not key:
            continue
        for column, aliases in _HEADER_ALIASES.items():
            if column not in columns and key in aliases:
                columns[column] = index
    return columns


def _name_queries_from_rows(rows: list[list], sheet_name: str = "") -> list[NameQuery]:
    if not rows:
        return []
    columns = _detect_name_columns(rows[0])
    name_index = columns.get("name")
    if name_index is None:
        return []

    def cell(row: list, column: str) -> str:
        index = columns.get(column)
        if index is None or index >= len(row):
            return ""
        return _clean(row[index])

    queries: list[NameQuery] = []
    for row_number, row in enumerate(rows[1:], start=2):
        company_name = cell(row, "name")
        if not company_name:
            continue
        queries.append(
            NameQuery(
                company_name=company_name,
                website=cell(row, "website"),
                email=cell(row, "email"),
                responsible_hint=cell(row, "responsible"),
                sheet_name=sheet_name,
                row_number=row_number,
            )
        )
    return queries


def _csv_rows(text: str) -> list[list]:
    return [row for row in csv.reader(io.StringIO(text))]


def parse_upload(filename: str, data: bytes) -> list[str]:
    return parse_upload_details(filename, data).cnpjs


def parse_upload_details(filename: str, data: bytes) -> UploadDetails:
    suffix = Path(filename).suffix.lower()
    if suffix in {".txt", ".csv"}:
        text = _decode(data)
        cnpjs = extract_cnpjs(text)
        source_type = suffix.lstrip(".")
        if cnpjs:
            return UploadDetails(filename=filename, cnpjs=cnpjs, source_type=source_type, mode="cnpj")
        name_queries = _name_queries_from_rows(_csv_rows(text)) if suffix == ".csv" else []
        if name_queries:
            return UploadDetails(
                filename=filename,
                cnpjs=[],
                source_type=source_type,
                mode="name",
                name_queries=name_queries,
            )
        return UploadDetails(filename=filename, cnpjs=[], source_type=source_type, mode="cnpj")

    if suffix == ".xlsx":
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        cnpjs: list[str] = []
        row_refs: list[dict] = []
        sheet_rows: list[tuple[str, list[list]]] = []
        for sheet in workbook.worksheets:
            rows: list[list] = []
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = ["" if value is None else str(value) for value in row]
                rows.append(values)
                row_cnpjs = extract_cnpjs("\n".join(values))
                cnpjs.extend(row_cnpjs)
                if row_cnpjs:
                    row_refs.append(
                        {
                            "sheet_name": sheet.title,
                            "row_number": row_number,
                            "cnpjs": row_cnpjs,
                        }
                    )
            sheet_rows.append((sheet.title, rows))
        workbook.close()
        if cnpjs:
            return UploadDetails(
                filename=filename,
                cnpjs=cnpjs,
                row_refs=row_refs,
                source_type="xlsx",
                mode="cnpj",
            )
        name_queries: list[NameQuery] = []
        for sheet_name, rows in sheet_rows:
            name_queries.extend(_name_queries_from_rows(rows, sheet_name=sheet_name))
        if name_queries:
            return UploadDetails(
                filename=filename,
                cnpjs=[],
                source_type="xlsx",
                mode="name",
                name_queries=name_queries,
            )
        return UploadDetails(filename=filename, cnpjs=[], source_type="xlsx", mode="cnpj")

    raise ValueError("Unsupported file type. Use .txt, .csv, or .xlsx.")
