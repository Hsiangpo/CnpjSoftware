from __future__ import annotations

import hashlib
import json
import threading
import time
import uuid
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from .cnpj import dedupe_preserve_order, normalize_cnpj
from .models import BatchResult, is_business_success


@dataclass
class ResumeState:
    upload_id: str
    done_count: int
    remaining_count: int
    total_count: int
    completed: bool
    updated_at: float

    def to_dict(self) -> dict:
        return {
            "upload_id": self.upload_id,
            "done_count": self.done_count,
            "remaining_count": self.remaining_count,
            "total_count": self.total_count,
            "completed": self.completed,
            "updated_at": self.updated_at,
        }


def _is_business_success_payload(data: dict | None) -> bool:
    payload = data or {}
    status = str(payload.get("status", ""))
    if status == "success":
        return True
    if status != "partial_success":
        return False
    responsible = payload.get("responsible") or {}
    names = responsible.get("names") or []
    return bool(
        responsible.get("analysis_source") == "rule_fallback"
        and any(str(name or "").strip() for name in names)
    )


class CheckpointStore:
    def __init__(self, root: Path) -> None:
        self.root = root
        self.root.mkdir(parents=True, exist_ok=True)
        self._lock = threading.Lock()

    def build_upload_id(self, filename: str, data: bytes, cnpjs: list[str]) -> str:
        digest = hashlib.sha256()
        digest.update(filename.encode("utf-8", errors="ignore"))
        digest.update(b"\0")
        digest.update(data)
        digest.update(b"\0")
        digest.update("\n".join(dedupe_preserve_order(cnpjs)).encode("utf-8"))
        return digest.hexdigest()

    def _path(self, upload_id: str) -> Path:
        return self.root / f"{upload_id}.json"

    def _source_path(self, upload_id: str, filename: str) -> Path:
        suffix = Path(filename).suffix or ".bin"
        return self.root / f"{upload_id}{suffix}"

    def _load_payload(self, upload_id: str) -> dict | None:
        path = self._path(upload_id)
        if not path.exists():
            return None
        return json.loads(path.read_text(encoding="utf-8"))

    def _write_text_atomic(self, path: Path, content: str) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        temp_path = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
        try:
            temp_path.write_text(content, encoding="utf-8")
            temp_path.replace(path)
        finally:
            if temp_path.exists():
                temp_path.unlink()

    def find_registered_upload(self, *, filename: str, data: bytes) -> dict | None:
        newest: dict | None = None
        newest_updated_at = -1.0
        for path in self.root.glob("*.json"):
            try:
                payload = json.loads(path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                continue
            if payload.get("filename") != filename:
                continue
            source_path = Path(str(payload.get("source_path", "")))
            try:
                if not source_path.exists() or source_path.read_bytes() != data:
                    continue
            except OSError:
                continue
            updated_at = float(payload.get("updated_at", 0.0))
            if updated_at > newest_updated_at:
                newest = payload
                newest_updated_at = updated_at
        return newest

    def register_upload(
        self,
        *,
        upload_id: str,
        filename: str,
        data: bytes,
        input_cnpjs: list[str],
        row_refs: list[dict],
        source_type: str,
    ) -> None:
        payload = self._load_payload(upload_id) or {
            "upload_id": upload_id,
            "results": {},
            "updated_at": 0.0,
        }
        payload["filename"] = filename
        payload["raw_input_cnpjs"] = [normalize_cnpj(item) for item in input_cnpjs]
        payload["input_cnpjs"] = dedupe_preserve_order([normalize_cnpj(item) for item in input_cnpjs])
        payload["row_refs"] = row_refs
        payload["source_type"] = source_type
        payload["source_path"] = str(self._source_path(upload_id, filename))
        payload["updated_at"] = time.time()
        with self._lock:
            self._source_path(upload_id, filename).write_bytes(data)
            self._write_text_atomic(self._path(upload_id), json.dumps(payload, ensure_ascii=False, indent=2))

    def get_resume_state(self, upload_id: str) -> ResumeState:
        payload = self._load_payload(upload_id)
        return self._resume_state_from_payload(upload_id, payload)

    def _resume_state_from_payload(self, upload_id: str, payload: dict | None) -> ResumeState:
        if not payload:
            return ResumeState(
                upload_id=upload_id,
                done_count=0,
                remaining_count=0,
                total_count=0,
                completed=False,
                updated_at=0.0,
            )
        results = payload.get("results", {})
        row_refs = payload.get("row_refs", [])
        if row_refs:
            total = len(row_refs)
            done_count = 0
            for item in row_refs:
                row_cnpjs = [normalize_cnpj(cnpj) for cnpj in item.get("cnpjs", [])]
                if row_cnpjs and all(cnpj in results for cnpj in row_cnpjs):
                    done_count += 1
        else:
            total = len(payload.get("input_cnpjs", []))
            done_count = len(results)
        return ResumeState(
            upload_id=upload_id,
            done_count=done_count,
            remaining_count=max(0, total - done_count),
            total_count=total,
            completed=done_count == total and total > 0,
            updated_at=float(payload.get("updated_at", 0.0)),
        )

    def get_progress_summary(self, upload_id: str) -> tuple[ResumeState, int, int]:
        payload = self._load_payload(upload_id)
        resume = self._resume_state_from_payload(upload_id, payload)
        if not payload:
            return resume, 0, 0
        stored = payload.get("results", {})
        normal_count = sum(1 for item in stored.values() if _is_business_success_payload(item))
        abnormal_count = max(0, len(stored) - normal_count)
        return resume, normal_count, abnormal_count

    def load_results(self, upload_id: str) -> list[BatchResult]:
        payload = self._load_payload(upload_id)
        if not payload:
            return []
        stored = payload.get("results", {})
        return [
            BatchResult.from_dict(stored[key])
            for key in payload.get("input_cnpjs", [])
            if key in stored
        ]

    def load_upload_metadata(self, upload_id: str) -> dict | None:
        return self._load_payload(upload_id)

    def save_results(
        self,
        *,
        upload_id: str,
        filename: str,
        input_cnpjs: list[str],
        results: list[BatchResult],
    ) -> None:
        normalized = dedupe_preserve_order([normalize_cnpj(item) for item in input_cnpjs])
        payload = {
            "upload_id": upload_id,
            "filename": filename,
            "input_cnpjs": normalized,
            "results": {result.normalized_cnpj: result.to_dict() for result in results},
            "updated_at": time.time(),
        }
        path = self._path(upload_id)
        with self._lock:
            self._write_text_atomic(path, json.dumps(payload, ensure_ascii=False, indent=2))

    def upsert_result(
        self,
        *,
        upload_id: str,
        filename: str,
        input_cnpjs: list[str],
        result: BatchResult,
    ) -> None:
        normalized = dedupe_preserve_order([normalize_cnpj(item) for item in input_cnpjs])
        with self._lock:
            payload = self._load_payload(upload_id) or {
                "upload_id": upload_id,
                "filename": filename,
                "input_cnpjs": normalized,
                "results": {},
                "updated_at": 0.0,
            }
            payload["filename"] = filename
            payload["input_cnpjs"] = normalized
            payload["results"][result.normalized_cnpj] = result.to_dict()
            payload["updated_at"] = time.time()
            self._write_text_atomic(self._path(upload_id), json.dumps(payload, ensure_ascii=False, indent=2))

    def build_enriched_xlsx(self, *, upload_id: str, results: list[BatchResult]) -> bytes:
        payload = self._load_payload(upload_id)
        if not payload or payload.get("source_type") != "xlsx":
            raise FileNotFoundError(upload_id)
        source_path = Path(payload["source_path"])
        workbook = load_workbook(source_path)
        result_map = {result.normalized_cnpj: result for result in results}
        headers = [
            "负责人姓名",
            "负责人角色",
            "状态",
            "分析来源",
            "模型",
        ]
        row_ref_map = {}
        for item in payload.get("row_refs", []):
            row_ref_map[(item["sheet_name"], item["row_number"])] = item.get("cnpjs", [])

        for sheet in workbook.worksheets:
            max_col = sheet.max_column
            for offset, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=max_col + offset, value=header)
            for row_number in range(2, sheet.max_row + 1):
                row_cnpjs = row_ref_map.get((sheet.title, row_number), [])
                matched_results = [result_map.get(normalize_cnpj(cnpj)) for cnpj in row_cnpjs]
                matched_results = [item for item in matched_results if item]
                names = "; ".join(
                    "; ".join(result.responsible.names) for result in matched_results if result.responsible
                )
                role = "; ".join(
                    result.responsible.role for result in matched_results if result.responsible and result.responsible.role
                )
                status = "; ".join("success" if is_business_success(result) else result.status for result in matched_results)
                source = "; ".join(
                    result.responsible.analysis_source for result in matched_results if result.responsible
                )
                model = "; ".join(
                    result.responsible.model_used for result in matched_results if result.responsible and result.responsible.model_used
                )
                values = [names, role, status, source, model]
                for offset, value in enumerate(values, start=1):
                    sheet.cell(row=row_number, column=max_col + offset, value=value)

        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()
        return buffer.getvalue()

    def build_summary_csv(self, *, results: list[BatchResult]) -> bytes:
        buffer = BytesIO()
        rows = [
            [
                "CNPJ",
                "Company",
                "Responsible",
                "Role",
                "Status",
                "AnalysisSource",
                "ModelUsed",
                "URL",
            ]
        ]
        for item in results:
            company = item.company or None
            responsible = item.responsible or None
            rows.append(
                [
                    item.input_cnpj or item.normalized_cnpj or "",
                    (company.trade_name or company.legal_name) if company else "",
                    "; ".join(responsible.names) if responsible else "",
                    responsible.role if responsible else "",
                    "success" if is_business_success(item) else item.status,
                    responsible.analysis_source if responsible else "",
                    responsible.model_used if responsible else "",
                    (company.url if company else "") or f"https://cnpj.biz/{item.normalized_cnpj}",
                ]
            )
        text_buffer = []
        for row in rows:
            csv_row = []
            for value in row:
                text = str(value).replace('"', '""')
                csv_row.append(f'"{text}"')
                text_buffer.append(",".join(csv_row))
        return ("\n".join(text_buffer) + "\n").encode("utf-8")

    def _write_bytes_atomic(self, path: Path, content: bytes) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        temp_path = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
        try:
            temp_path.write_bytes(content)
            temp_path.replace(path)
        finally:
            if temp_path.exists():
                temp_path.unlink()

    def materialize_output(
        self,
        *,
        upload_id: str,
        filename: str,
        output_path: Path,
        results: list[BatchResult],
    ) -> Path:
        payload = self._load_payload(upload_id)
        if not payload:
            raise FileNotFoundError(upload_id)
        source_type = payload.get("source_type", "")
        if source_type == "xlsx":
            content = self.build_enriched_xlsx(upload_id=upload_id, results=results)
            self._write_bytes_atomic(output_path, content)
            return output_path
        self._write_bytes_atomic(output_path, self.build_summary_csv(results=results))
        return output_path
