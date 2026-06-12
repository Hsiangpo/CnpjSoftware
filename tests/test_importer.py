from openpyxl import Workbook

from cnpj_tool.importer import parse_upload, parse_upload_details


def test_parse_upload_reads_text_and_csv_bytes():
    assert parse_upload("cnpjs.txt", b"03.541.629/0001-37\n21.746.991/0001-26") == [
        "03541629000137",
        "21746991000126",
    ]

    assert parse_upload("cnpjs.csv", "nome,cnpj\nA,02.759.853/0001-37".encode()) == [
        "02759853000137",
    ]


def test_parse_upload_reads_xlsx_bytes(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "empresa"
    sheet["B2"] = "00.642.475/0001-81"
    path = tmp_path / "cnpjs.xlsx"
    workbook.save(path)

    assert parse_upload("cnpjs.xlsx", path.read_bytes()) == ["00642475000181"]


def test_parse_upload_details_tracks_xlsx_rows_and_duplicates(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "batch1"
    sheet.append(["姓名", "CNPJ", "公司"])
    sheet.append(["Lucas", "00.001.180/0001-26", "Eletrobras"])
    sheet.append(["Andre", "00.001.180/0001-26", "Eletrobras"])
    sheet.append(["Fabio", "00.003.699/0001-43", "BASA"])
    path = tmp_path / "cnpjs.xlsx"
    workbook.save(path)

    details = parse_upload_details("cnpjs.xlsx", path.read_bytes())

    assert details.cnpjs == [
        "00001180000126",
        "00001180000126",
        "00003699000143",
    ]
    assert details.filename == "cnpjs.xlsx"
    assert details.row_refs[0]["sheet_name"] == "batch1"
    assert details.row_refs[0]["row_number"] == 2
    assert details.row_refs[0]["cnpjs"] == ["00001180000126"]
    assert details.row_refs[1]["row_number"] == 3


def test_cnpj_csv_is_still_detected_as_cnpj_mode():
    details = parse_upload_details("cnpjs.csv", "nome,cnpj\nA,02.759.853/0001-37".encode())
    assert details.mode == "cnpj"
    assert details.cnpjs == ["02759853000137"]
    assert details.name_queries == []


def test_parse_upload_details_reads_gbk_company_name_csv():
    header = "法人,公司名称,公司网址,邮箱地址"
    body = [
        "Rita de Cassia Yazbek,Construtora R.Yazbek LTDA,http://www.ryazbek.com.br,ana@ryazbek.com.br",
        "Douglas do Vale Santiago,Aerotrafic Transportes e Logística,https://aerotrafic.com.br,caroline.almeida@aerotrafic.com.br",
    ]
    data = ("\n".join([header] + body)).encode("gb18030")

    details = parse_upload_details("巴西最新到130.csv", data)

    assert details.mode == "name"
    assert details.cnpjs == []
    assert len(details.name_queries) == 2
    first = details.name_queries[0]
    assert first.company_name == "Construtora R.Yazbek LTDA"
    assert first.website == "http://www.ryazbek.com.br"
    assert first.email == "ana@ryazbek.com.br"
    assert first.responsible_hint == "Rita de Cassia Yazbek"
    assert first.row_number == 2
    assert details.name_queries[1].company_name == "Aerotrafic Transportes e Logística"


def test_parse_upload_details_reads_company_name_xlsx(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "dados"
    sheet.append(["法人", "公司名称", "邮箱地址"])
    sheet.append(["Armando Masao Abe", "Transkompa", "daniel.burguetti@transkompa.com.br"])
    path = tmp_path / "names.xlsx"
    workbook.save(path)

    details = parse_upload_details("names.xlsx", path.read_bytes())

    assert details.mode == "name"
    assert len(details.name_queries) == 1
    query = details.name_queries[0]
    assert query.company_name == "Transkompa"
    assert query.email == "daniel.burguetti@transkompa.com.br"
    assert query.sheet_name == "dados"
    assert query.row_number == 2
