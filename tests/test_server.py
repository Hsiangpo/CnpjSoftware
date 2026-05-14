from pathlib import Path

import cnpj_tool.server as server_module
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from cnpj_tool.models import BatchResult, CompanyData, ResponsibleResult
from cnpj_tool.server import create_app


def test_health_endpoint_reports_ready(tmp_path, monkeypatch):
    env_path = tmp_path / ".env"
    env_path.write_text("", encoding="utf-8")
    monkeypatch.setenv("CNPJ_TOOL_ENV_FILE", str(env_path))
    client = TestClient(create_app())

    response = client.get("/api/health")

    assert response.status_code == 200
    assert response.json()["status"] == "ok"
    assert response.json()["provider_order"] == ["brasilapi", "cnpjbiz"]
    assert response.json()["fallback_models"] == []
    assert "browser_proxy" in response.json()


def test_source_files_endpoint_lists_root_cnpj_directory(tmp_path, monkeypatch):
    input_dir = tmp_path / "cnpj"
    output_dir = tmp_path / "output"
    checkpoint_dir = tmp_path / "checkpoints"
    input_dir.mkdir()
    output_dir.mkdir()
    checkpoint_dir.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "batch1"
    sheet.append(["公司名称", "CNPJ"])
    sheet.append(["Empresa Teste", "03.541.629/0001-37"])
    workbook.save(input_dir / "sample.xlsx")
    workbook.close()
    output_file = output_dir / "sample-responsaveis.xlsx"
    output_file.write_bytes(b"artifact")
    monkeypatch.setenv("CNPJ_TOOL_INPUT_DIR", str(input_dir))
    monkeypatch.setenv("CNPJ_TOOL_OUTPUT_DIR", str(output_dir))
    monkeypatch.setenv("CNPJ_TOOL_CHECKPOINT_DIR", str(checkpoint_dir))

    client = TestClient(create_app(auto_run_jobs=False))
    response = client.get("/api/source-files")

    assert response.status_code == 200
    payload = response.json()
    assert [item["name"] for item in payload["files"]] == ["sample.xlsx"]
    assert payload["files"][0]["source_type"] == "xlsx"
    assert payload["files"][0]["count"] == 1
    assert payload["files"][0]["unique_count"] == 1
    assert payload["files"][0]["resume"]["total_count"] == 0
    assert "path" not in payload["files"][0]
    assert payload["files"][0]["output_name"] == "sample-responsaveis.xlsx"
    assert payload["files"][0]["output_exists"] is True
    assert payload["files"][0]["output_size_bytes"] == 8
    assert payload["files"][0]["normal_count"] == 0
    assert payload["files"][0]["abnormal_count"] == 0


def test_output_files_endpoint_lists_recent_artifacts(tmp_path, monkeypatch):
    input_dir = tmp_path / "cnpj"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    output_dir.mkdir()
    newer = output_dir / "b-responsaveis.xlsx"
    older = output_dir / "a-responsaveis.csv"
    older.write_text("a", encoding="utf-8")
    newer.write_text("b", encoding="utf-8")
    older.touch()
    newer.touch()
    monkeypatch.setenv("CNPJ_TOOL_INPUT_DIR", str(input_dir))
    monkeypatch.setenv("CNPJ_TOOL_OUTPUT_DIR", str(output_dir))

    client = TestClient(create_app(auto_run_jobs=False))
    response = client.get("/api/output-files")

    assert response.status_code == 200
    payload = response.json()
    assert payload["output_dir"] == str(output_dir)
    assert [item["name"] for item in payload["files"]] == ["b-responsaveis.xlsx", "a-responsaveis.csv"]
    assert payload["files"][0]["size_bytes"] == 1


def test_open_output_directory_endpoint_uses_platform_open_command(tmp_path, monkeypatch):
    input_dir = tmp_path / "cnpj"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    output_dir.mkdir()
    monkeypatch.setenv("CNPJ_TOOL_INPUT_DIR", str(input_dir))
    monkeypatch.setenv("CNPJ_TOOL_OUTPUT_DIR", str(output_dir))
    launches = []

    def fake_run(cmd, **kwargs):
        launches.append(cmd)

        class Result:
            returncode = 0
            stderr = ""
            stdout = ""

        return Result()

    monkeypatch.setattr(server_module.subprocess, "run", fake_run)
    monkeypatch.setattr(server_module.shutil, "which", lambda name: "/usr/bin/xdg-open" if name == "xdg-open" else None)
    monkeypatch.setattr(server_module.platform, "system", lambda: "Linux")

    client = TestClient(create_app(auto_run_jobs=False))
    response = client.post("/api/output-directory/open")

    assert response.status_code == 200
    assert response.json()["opened"] is True
    assert launches == [["xdg-open", str(output_dir)]]


def test_open_output_directory_endpoint_uses_windows_explorer(tmp_path, monkeypatch):
    input_dir = tmp_path / "cnpj"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    output_dir.mkdir()
    monkeypatch.setenv("CNPJ_TOOL_INPUT_DIR", str(input_dir))
    monkeypatch.setenv("CNPJ_TOOL_OUTPUT_DIR", str(output_dir))
    launches = []

    def fake_popen(cmd, **kwargs):
        launches.append(cmd)

        class Dummy:
            pass

        return Dummy()

    monkeypatch.setattr(server_module.subprocess, "Popen", fake_popen)
    monkeypatch.setattr(server_module.platform, "system", lambda: "Windows")

    client = TestClient(create_app(auto_run_jobs=False))
    response = client.post("/api/output-directory/open")

    assert response.status_code == 200
    assert launches == [["explorer", str(output_dir)]]


def test_open_output_directory_endpoint_uses_macos_open(tmp_path, monkeypatch):
    input_dir = tmp_path / "cnpj"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    output_dir.mkdir()
    monkeypatch.setenv("CNPJ_TOOL_INPUT_DIR", str(input_dir))
    monkeypatch.setenv("CNPJ_TOOL_OUTPUT_DIR", str(output_dir))
    launches = []

    def fake_run(cmd, **kwargs):
        launches.append(cmd)

        class Result:
            returncode = 0
            stderr = ""
            stdout = ""

        return Result()

    monkeypatch.setattr(server_module.subprocess, "run", fake_run)
    monkeypatch.setattr(server_module.platform, "system", lambda: "Darwin")

    client = TestClient(create_app(auto_run_jobs=False))
    response = client.post("/api/output-directory/open")

    assert response.status_code == 200
    assert launches == [["open", str(output_dir)]]


def test_open_output_file_endpoint_uses_platform_open_command(tmp_path, monkeypatch):
    input_dir = tmp_path / "cnpj"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    output_dir.mkdir()
    artifact = output_dir / "sample-responsaveis.xlsx"
    artifact.write_text("artifact", encoding="utf-8")
    monkeypatch.setenv("CNPJ_TOOL_INPUT_DIR", str(input_dir))
    monkeypatch.setenv("CNPJ_TOOL_OUTPUT_DIR", str(output_dir))
    launches = []

    def fake_run(cmd, **kwargs):
        launches.append(cmd)

        class Result:
            returncode = 0
            stderr = ""
            stdout = ""

        return Result()

    monkeypatch.setattr(server_module.subprocess, "run", fake_run)
    monkeypatch.setattr(server_module.shutil, "which", lambda name: "/usr/bin/xdg-open" if name == "xdg-open" else None)
    monkeypatch.setattr(server_module.platform, "system", lambda: "Linux")

    client = TestClient(create_app(auto_run_jobs=False))
    response = client.post("/api/output-files/sample-responsaveis.xlsx/open")

    assert response.status_code == 200
    assert response.json()["opened"] is True
    assert launches == [["xdg-open", str(artifact)]]


def test_open_output_file_endpoint_surfaces_launcher_failure(tmp_path, monkeypatch):
    input_dir = tmp_path / "cnpj"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    output_dir.mkdir()
    artifact = output_dir / "sample-responsaveis.xlsx"
    artifact.write_text("artifact", encoding="utf-8")
    monkeypatch.setenv("CNPJ_TOOL_INPUT_DIR", str(input_dir))
    monkeypatch.setenv("CNPJ_TOOL_OUTPUT_DIR", str(output_dir))

    def fake_run(cmd, **kwargs):
        class Result:
            returncode = 4
            stderr = "No application is registered as handling this file"
            stdout = ""

        return Result()

    monkeypatch.setattr(server_module.subprocess, "run", fake_run)
    monkeypatch.setattr(server_module.shutil, "which", lambda name: "/usr/bin/xdg-open" if name == "xdg-open" else None)
    monkeypatch.setattr(server_module.platform, "system", lambda: "Linux")

    client = TestClient(create_app(auto_run_jobs=False))
    response = client.post("/api/output-files/sample-responsaveis.xlsx/open")

    assert response.status_code == 500
    assert "No application is registered" in response.json()["detail"]


def test_directory_backed_job_writes_enriched_output_file(tmp_path, monkeypatch):
    input_dir = tmp_path / "cnpj"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    output_dir.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "batch1"
    sheet.append(["公司名称", "CNPJ"])
    sheet.append(["Empresa Teste", "03.541.629/0001-37"])
    source_path = input_dir / "sample.xlsx"
    workbook.save(source_path)
    workbook.close()
    monkeypatch.setenv("CNPJ_TOOL_INPUT_DIR", str(input_dir))
    monkeypatch.setenv("CNPJ_TOOL_OUTPUT_DIR", str(output_dir))

    class FakeAnalyzer:
        def analyze_many(self, cnpjs, existing_results=None, on_result=None, should_stop=None):
            result = BatchResult(
                input_cnpj="03.541.629/0001-37",
                normalized_cnpj="03541629000137",
                status="success",
                company=CompanyData(
                    cnpj="03541629000137",
                    formatted_cnpj="03.541.629/0001-37",
                    url="https://cnpj.biz/03541629000137",
                    legal_name="Empresa Teste LTDA",
                    city="Belo Horizonte",
                    state="MG",
                ),
                responsible=ResponsibleResult(
                    names=["Maria Teste"],
                    role="Socio-Administrador",
                    confidence=0.91,
                    reasoning="rule",
                    analysis_source="rule_fallback",
                ),
            )
            if on_result:
                on_result(result)
            return [result]

    monkeypatch.setattr(server_module, "build_analyzer", lambda: FakeAnalyzer())

    client = TestClient(create_app())
    created = client.post("/api/jobs", json={"source_name": "sample.xlsx"})

    assert created.status_code == 200
    payload = created.json()
    assert payload["filename"] == "sample.xlsx"
    assert payload["output_path"].endswith("sample-responsaveis.xlsx")

    job = client.get(f"/api/jobs/{payload['job_id']}").json()
    assert job["status"] == "completed"
    assert Path(job["output_path"]).exists()

    output_workbook = load_workbook(job["output_path"])
    output_sheet = output_workbook["batch1"]
    headers = [output_sheet.cell(row=1, column=index).value for index in range(1, output_sheet.max_column + 1)]
    row_values = [output_sheet.cell(row=2, column=index).value for index in range(1, output_sheet.max_column + 1)]
    output_workbook.close()

    assert "负责人姓名" in headers
    assert "置信度" not in headers
    assert "依据" not in headers
    assert "Provider Trace" not in headers
    assert "Maria Teste" in row_values


def test_retry_failed_endpoint_creates_retry_job_with_failed_cnpjs(tmp_path, monkeypatch):
    input_dir = tmp_path / "cnpj"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    output_dir.mkdir()
    monkeypatch.setenv("CNPJ_TOOL_INPUT_DIR", str(input_dir))
    monkeypatch.setenv("CNPJ_TOOL_OUTPUT_DIR", str(output_dir))

    app = create_app(auto_run_jobs=False)
    failed = BatchResult(
        input_cnpj="03.541.629/0001-37",
        normalized_cnpj="03541629000137",
        status="fetch_error",
        error="timeout",
    )
    success = BatchResult(
        input_cnpj="21.746.991/0001-26",
        normalized_cnpj="21746991000126",
        status="success",
        company=CompanyData(
            cnpj="21746991000126",
            formatted_cnpj="21.746.991/0001-26",
            url="https://cnpj.biz/21746991000126",
            legal_name="Empresa Ok LTDA",
        ),
    )
    original = app.state.jobs.create(
        ["03541629000137", "21746991000126"],
        filename="sample.xlsx",
        source_name="sample.xlsx",
        existing_results=[failed, success],
    )
    client = TestClient(app)

    response = client.post(f"/api/jobs/{original.job_id}/retry-failed")

    assert response.status_code == 200
    retry_job = response.json()
    assert retry_job["input_cnpjs"] == ["03541629000137"]
    assert retry_job["filename"] == "sample-failed-retry.csv"
    assert retry_job["output_path"].endswith("sample-failed-retry.csv")


def test_retry_failed_endpoint_includes_partial_success_and_dedupes(tmp_path, monkeypatch):
    input_dir = tmp_path / "cnpj"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    output_dir.mkdir()
    monkeypatch.setenv("CNPJ_TOOL_INPUT_DIR", str(input_dir))
    monkeypatch.setenv("CNPJ_TOOL_OUTPUT_DIR", str(output_dir))

    app = create_app(auto_run_jobs=False)
    partial = BatchResult(
        input_cnpj="03.541.629/0001-37",
        normalized_cnpj="03541629000137",
        status="partial_success",
        error="llm fallback",
    )
    duplicate_partial = BatchResult(
        input_cnpj="03.541.629/0001-37",
        normalized_cnpj="03541629000137",
        status="partial_success",
        error="llm fallback",
    )
    failed = BatchResult(
        input_cnpj="21.746.991/0001-26",
        normalized_cnpj="21746991000126",
        status="fetch_error",
        error="timeout",
    )
    success = BatchResult(
        input_cnpj="00.000.000/0001-91",
        normalized_cnpj="00000000000191",
        status="success",
        company=CompanyData(
            cnpj="00000000000191",
            formatted_cnpj="00.000.000/0001-91",
            url="https://cnpj.biz/00000000000191",
            legal_name="Empresa Ok LTDA",
        ),
    )
    original = app.state.jobs.create(
        ["03541629000137", "03541629000137", "21746991000126", "00000000000191"],
        filename="sample.xlsx",
        source_name="sample.xlsx",
        existing_results=[partial, duplicate_partial, failed, success],
    )
    client = TestClient(app)

    response = client.post(f"/api/jobs/{original.job_id}/retry-failed")

    assert response.status_code == 200
    retry_job = response.json()
    assert retry_job["input_cnpjs"] == ["03541629000137", "21746991000126"]


def test_retry_failed_endpoint_reuses_original_source_output_for_directory_jobs(tmp_path, monkeypatch):
    input_dir = tmp_path / "cnpj"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    output_dir.mkdir()
    monkeypatch.setenv("CNPJ_TOOL_INPUT_DIR", str(input_dir))
    monkeypatch.setenv("CNPJ_TOOL_OUTPUT_DIR", str(output_dir))

    app = create_app(auto_run_jobs=False)
    original = app.state.jobs.create(
        ["03541629000137", "21746991000126"],
        upload_id="upload-1",
        source_name="sample.xlsx",
        filename="sample.xlsx",
        output_path=str(output_dir / "sample-responsaveis.xlsx"),
        existing_results=[
            BatchResult(
                input_cnpj="03.541.629/0001-37",
                normalized_cnpj="03541629000137",
                status="fetch_error",
                error="timeout",
            )
            ],
        )

    class FakeStore:
        def load_results(self, upload_id):
            assert upload_id == "upload-1"
            return [
                BatchResult(
                    input_cnpj="03.541.629/0001-37",
                    normalized_cnpj="03541629000137",
                    status="fetch_error",
                    error="timeout",
                )
            ]

    monkeypatch.setattr(server_module, "get_checkpoint_store", lambda _app: FakeStore())
    client = TestClient(app)

    response = client.post(f"/api/jobs/{original.job_id}/retry-failed")

    assert response.status_code == 200
    retry_job = response.json()
    assert retry_job["upload_id"] == "upload-1"
    assert retry_job["filename"] == "sample.xlsx"
    assert retry_job["output_path"].endswith("sample-responsaveis.xlsx")


def test_retry_one_endpoint_uses_original_source_context(tmp_path, monkeypatch):
    input_dir = tmp_path / "cnpj"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    output_dir.mkdir()
    monkeypatch.setenv("CNPJ_TOOL_INPUT_DIR", str(input_dir))
    monkeypatch.setenv("CNPJ_TOOL_OUTPUT_DIR", str(output_dir))

    app = create_app(auto_run_jobs=False)
    original = app.state.jobs.create(
        ["03541629000137", "21746991000126"],
        upload_id="upload-1",
        source_name="sample.xlsx",
        filename="sample.xlsx",
        output_path=str(output_dir / "sample-responsaveis.xlsx"),
        existing_results=[
            BatchResult(
                input_cnpj="03.541.629/0001-37",
                normalized_cnpj="03541629000137",
                status="fetch_error",
                error="timeout",
            ),
            BatchResult(
                input_cnpj="21.746.991/0001-26",
                normalized_cnpj="21746991000126",
                status="success",
            ),
        ],
    )
    client = TestClient(app)

    response = client.post(f"/api/jobs/{original.job_id}/retry-one", json={"cnpj": "03.541.629/0001-37"})

    assert response.status_code == 200
    retry_job = response.json()
    assert retry_job["input_cnpjs"] == ["03541629000137"]
    assert retry_job["upload_id"] == "upload-1"
    assert retry_job["filename"] == "sample.xlsx"
    assert retry_job["output_path"].endswith("sample-responsaveis.xlsx")


def test_retry_failed_endpoint_uses_checkpoint_results_for_source_jobs(tmp_path, monkeypatch):
    input_dir = tmp_path / "cnpj"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    output_dir.mkdir()
    monkeypatch.setenv("CNPJ_TOOL_INPUT_DIR", str(input_dir))
    monkeypatch.setenv("CNPJ_TOOL_OUTPUT_DIR", str(output_dir))

    app = create_app(auto_run_jobs=False)
    original = app.state.jobs.create(
        ["03541629000137"],
        upload_id="upload-1",
        source_name="sample.xlsx",
        filename="sample.xlsx",
        output_path=str(output_dir / "sample-responsaveis.xlsx"),
        existing_results=[
            BatchResult(
                input_cnpj="03.541.629/0001-37",
                normalized_cnpj="03541629000137",
                status="fetch_error",
                error="stale timeout",
            )
        ],
    )

    class FakeStore:
        def load_results(self, upload_id):
            assert upload_id == "upload-1"
            return [
                BatchResult(
                    input_cnpj="03.541.629/0001-37",
                    normalized_cnpj="03541629000137",
                    status="success",
                )
            ]

    monkeypatch.setattr(server_module, "get_checkpoint_store", lambda _app: FakeStore())
    client = TestClient(app)

    response = client.post(f"/api/jobs/{original.job_id}/retry-failed")

    assert response.status_code == 400
    assert response.json()["detail"] == "This job has no failed CNPJ values"


def test_source_file_summary_clears_abnormal_after_successful_retry(tmp_path, monkeypatch):
    input_dir = tmp_path / "cnpj"
    output_dir = tmp_path / "output"
    checkpoint_dir = tmp_path / "checkpoints"
    input_dir.mkdir()
    output_dir.mkdir()
    checkpoint_dir.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "batch1"
    sheet.append(["公司名称", "CNPJ"])
    sheet.append(["Empresa Teste", "03.541.629/0001-37"])
    source_path = input_dir / "sample.xlsx"
    workbook.save(source_path)
    workbook.close()
    monkeypatch.setenv("CNPJ_TOOL_INPUT_DIR", str(input_dir))
    monkeypatch.setenv("CNPJ_TOOL_OUTPUT_DIR", str(output_dir))
    monkeypatch.setenv("CNPJ_TOOL_CHECKPOINT_DIR", str(checkpoint_dir))

    class FakeAnalyzer:
        def __init__(self):
            self.calls = 0

        def analyze_many(self, cnpjs, existing_results=None, on_result=None, should_stop=None):
            self.calls += 1
            status = "fetch_error" if self.calls == 1 else "success"
            result = BatchResult(
                input_cnpj="03.541.629/0001-37",
                normalized_cnpj="03541629000137",
                status=status,
                error="timeout" if status != "success" else "",
                company=CompanyData(
                    cnpj="03541629000137",
                    formatted_cnpj="03.541.629/0001-37",
                    url="https://cnpj.biz/03541629000137",
                    legal_name="Empresa Teste LTDA",
                ) if status == "success" else None,
                responsible=ResponsibleResult(
                    names=["Maria Teste"],
                    role="Socio-Administrador",
                    confidence=0.91,
                    reasoning="rule",
                    analysis_source="rule_fallback",
                ) if status == "success" else None,
            )
            if on_result:
                on_result(result)
            return [result]

    analyzer = FakeAnalyzer()
    monkeypatch.setattr(server_module, "build_analyzer", lambda: analyzer)

    client = TestClient(create_app())
    first = client.post("/api/jobs", json={"source_name": "sample.xlsx"})
    assert first.status_code == 200
    original_job_id = first.json()["job_id"]
    first_job = client.get(f"/api/jobs/{original_job_id}").json()
    assert first_job["status"] == "completed"

    summary_after_fail = client.get("/api/source-files").json()["files"][0]
    assert summary_after_fail["abnormal_count"] == 1
    assert summary_after_fail["normal_count"] == 0

    retry = client.post(f"/api/jobs/{original_job_id}/retry-failed")
    assert retry.status_code == 200
    retry_job_id = retry.json()["job_id"]
    retried_job = client.get(f"/api/jobs/{retry_job_id}").json()
    assert retried_job["status"] == "completed"

    summary_after_retry = client.get("/api/source-files").json()["files"][0]
    assert summary_after_retry["abnormal_count"] == 0
    assert summary_after_retry["normal_count"] == 1

    retry_again = client.post(f"/api/jobs/{original_job_id}/retry-failed")
    assert retry_again.status_code == 400
    assert retry_again.json()["detail"] == "This job has no failed CNPJ values"


def test_settings_endpoint_updates_env_backed_runtime_settings(tmp_path, monkeypatch):
    env_path = tmp_path / ".env"
    env_path.write_text(
        "\n".join(
            [
                "LLM_API_KEY=old-key",
                "LLM_MODEL=gpt-5.4-mini",
                "LLM_FALLBACK_MODELS=",
                "LLM_TIMEOUT_SECONDS=30",
                "CNPJ_PROVIDER_ORDER=brasilapi,cnpjbiz",
                "CNPJ_BIZ_REQUEST_DELAY_SECONDS=0",
                "CNPJ_BIZ_USER_AGENT=Mozilla/5.0 Chrome/146",
                "BLURPATH_PROXY_HOST=blurpath.net",
                "BLURPATH_PROXY_PORT=15121",
                "BLURPATH_PROXY_PORTS=",
                "BLURPATH_PROXY_USERNAME=acct",
                "BLURPATH_PROXY_PASSWORD=pass",
                "BLURPATH_PROXY_NODES=",
                "SYSTEM_CONCURRENCY=2",
            ]
        ),
        encoding="utf-8",
    )
    monkeypatch.setenv("CNPJ_TOOL_ENV_FILE", str(env_path))

    client = TestClient(create_app(auto_run_jobs=False))

    initial = client.get("/api/settings")
    assert initial.status_code == 200
    assert initial.json()["llm_api_key"] == "<set>"
    assert "old-key" not in initial.text
    assert initial.json()["system_concurrency"] == 2
    assert initial.json()["cnpj_biz_user_agent"] == "Mozilla/5.0 Chrome/146"
    assert initial.json()["blurpath_proxy_ports"] == [15121]

    updated = client.put(
        "/api/settings",
        json={
            "llm_api_key": "new-key",
            "llm_model": "gpt-5.4",
            "system_concurrency": 5,
            "blurpath_proxy_ports": [15121],
        },
    )

    assert updated.status_code == 200
    assert updated.json()["llm_api_key"] == "<set>"
    assert "new-key" not in updated.text
    assert updated.json()["llm_model"] == "gpt-5.4"
    assert updated.json()["system_concurrency"] == 5
    assert updated.json()["blurpath_proxy_ports"] == [15121]
    assert "LLM_API_KEY=new-key" in env_path.read_text(encoding="utf-8")
    assert "SYSTEM_CONCURRENCY=5" in env_path.read_text(encoding="utf-8")


def test_settings_endpoint_updates_active_blurpath_proxy_node_pool(tmp_path, monkeypatch):
    env_path = tmp_path / ".env"
    env_path.write_text(
        "\n".join(
            [
                "LLM_API_KEY=old-key",
                "LLM_MODEL=gpt-5.4-mini",
                "CNPJ_PROVIDER_ORDER=brasilapi,cnpjbiz",
                "BLURPATH_PROXY_HOST=",
                "BLURPATH_PROXY_USERNAME=",
                "BLURPATH_PROXY_PASSWORD=",
                "BLURPATH_PROXY_PORTS=15129,15121",
                "BLURPATH_PROXY_NODES=http://acct:secret-pass@blurpath.net:15129|http://acct:secret-pass@blurpath.net:15121",
            ]
        ),
        encoding="utf-8",
    )
    monkeypatch.setenv("CNPJ_TOOL_ENV_FILE", str(env_path))

    client = TestClient(create_app(auto_run_jobs=False))

    initial = client.get("/api/settings")
    assert initial.status_code == 200
    assert initial.json()["blurpath_available_proxy_ports"] == [15129, 15121]
    assert initial.json()["blurpath_proxy_ports"] == [15129, 15121]
    assert "secret-pass" not in initial.text

    updated = client.put("/api/settings", json={"blurpath_proxy_ports": [15129]})

    assert updated.status_code == 200
    assert updated.json()["blurpath_available_proxy_ports"] == [15129, 15121]
    assert updated.json()["blurpath_proxy_ports"] == [15129]
    assert "secret-pass" not in updated.text
    proxy = client.get("/api/health").json()["browser_proxy"]
    assert proxy["ports"] == [15129]
    assert proxy["node_count"] == 1


def test_settings_endpoint_updates_direct_blurpath_proxy_runtime_fields(tmp_path, monkeypatch):
    env_path = tmp_path / ".env"
    env_path.write_text(
        "\n".join(
            [
                "LLM_API_KEY=old-key",
                "LLM_MODEL=gpt-5.4-mini",
                "CNPJ_PROVIDER_ORDER=brasilapi,cnpjbiz",
                "BLURPATH_PROXY_HOST=old.blurpath.net",
                "BLURPATH_PROXY_PORT=15121",
                "BLURPATH_PROXY_PORTS=15121",
                "BLURPATH_PROXY_PROTOCOL=http",
                "BLURPATH_PROXY_USERNAME=acct-old",
                "BLURPATH_PROXY_PASSWORD=old-secret",
                "BLURPATH_PROXY_NODES=",
                "BLURPATH_PROXY_REGION=BR",
                "BLURPATH_PROXY_SESSION_TIME_MINUTES=10",
            ]
        ),
        encoding="utf-8",
    )
    monkeypatch.setenv("CNPJ_TOOL_ENV_FILE", str(env_path))

    client = TestClient(create_app(auto_run_jobs=False))

    initial = client.get("/api/settings")
    assert initial.status_code == 200
    assert initial.json()["blurpath_proxy_host"] == "old.blurpath.net"
    assert initial.json()["blurpath_proxy_username"] == "acct-old"
    assert initial.json()["blurpath_proxy_password"] == "<set>"
    assert initial.json()["blurpath_proxy_region"] == "BR"
    assert initial.json()["blurpath_proxy_protocol"] == "http"
    assert initial.json()["blurpath_proxy_session_time_minutes"] == 10
    assert "old-secret" not in initial.text

    updated = client.put(
        "/api/settings",
        json={
            "blurpath_proxy_host": "new.blurpath.net",
            "blurpath_proxy_ports": [15129, 15121],
            "blurpath_proxy_protocol": "socks5",
            "blurpath_proxy_username": "acct-new",
            "blurpath_proxy_password": "new-secret",
            "blurpath_proxy_region": "US",
            "blurpath_proxy_session_time_minutes": 20,
        },
    )

    assert updated.status_code == 200
    assert updated.json()["blurpath_proxy_host"] == "new.blurpath.net"
    assert updated.json()["blurpath_proxy_username"] == "acct-new"
    assert updated.json()["blurpath_proxy_password"] == "<set>"
    assert updated.json()["blurpath_proxy_region"] == "US"
    assert updated.json()["blurpath_proxy_protocol"] == "socks5"
    assert updated.json()["blurpath_proxy_session_time_minutes"] == 20
    assert "new-secret" not in updated.text
    content = env_path.read_text(encoding="utf-8")
    assert "BLURPATH_PROXY_HOST=new.blurpath.net" in content
    assert "BLURPATH_PROXY_PROTOCOL=socks5" in content
    assert "BLURPATH_PROXY_USERNAME=acct-new" in content
    assert "BLURPATH_PROXY_PASSWORD=new-secret" in content
    assert "BLURPATH_PROXY_REGION=US" in content
    assert "BLURPATH_PROXY_SESSION_TIME_MINUTES=20" in content


def test_proxy_preflight_reports_each_blurpath_port_without_secrets(tmp_path, monkeypatch):
    env_path = tmp_path / ".env"
    env_path.write_text(
        "\n".join(
            [
                "BLURPATH_PROXY_HOST=blurpath.net",
                "BLURPATH_PROXY_PORT=15121",
                "BLURPATH_PROXY_PORTS=15121,15129",
                "BLURPATH_PROXY_USERNAME=acct",
                "BLURPATH_PROXY_PASSWORD=secret-pass",
                "BLURPATH_PROXY_REGION=BR",
                "SYSTEM_CONCURRENCY=2",
            ]
        ),
        encoding="utf-8",
    )
    monkeypatch.setenv("CNPJ_TOOL_ENV_FILE", str(env_path))

    def fake_probe(config, **kwargs):
        return {
            "provider": "blurpath",
            "ok": config.port == 15129,
            "region": config.region,
            "port": config.port,
            "protocol": config.protocol,
            "status_code": 200 if config.port == 15129 else None,
            "ip": "177.0.0.1" if config.port == 15129 else "",
            "error": "" if config.port == 15129 else "secret-pass tunnel timeout",
        }

    monkeypatch.setattr(server_module, "probe_blurpath_proxy", fake_probe)
    client = TestClient(create_app(auto_run_jobs=False))

    response = client.get("/api/proxy-preflight")

    assert response.status_code == 200
    payload = response.json()
    assert payload["proxy_probe_ok"] is True
    assert [item["port"] for item in payload["proxy_probe_results"]] == [15121, 15129]
    assert "secret-pass" not in response.text


def test_proxy_preflight_skips_probe_when_blurpath_not_configured(tmp_path, monkeypatch):
    env_path = tmp_path / ".env"
    env_path.write_text("SYSTEM_CONCURRENCY=2\n", encoding="utf-8")
    monkeypatch.setenv("CNPJ_TOOL_ENV_FILE", str(env_path))
    client = TestClient(create_app(auto_run_jobs=False))

    response = client.get("/api/proxy-preflight")

    assert response.status_code == 200
    payload = response.json()
    assert payload["proxy_probe_ok"] is False
    assert payload["proxy_error"] == "Blurpath proxy is not configured"
    assert payload["proxy_probe_results"] == []


def test_homepage_renders_current_queue_ui():
    client = TestClient(create_app(auto_run_jobs=False))

    response = client.get("/")

    assert response.status_code == 200
    assert 'id="sourceFileSelect"' in response.text
    assert 'id="queueSourceButton"' in response.text
    assert 'id="runFailedButton"' in response.text
    assert "/static/lucide.min.js" in response.text
    assert "/api/proxy-preflight" not in response.text


def test_app_shutdown_closes_analyzer(tmp_path, monkeypatch):
    env_path = tmp_path / ".env"
    env_path.write_text("", encoding="utf-8")
    monkeypatch.setenv("CNPJ_TOOL_ENV_FILE", str(env_path))
    closed = []

    class FakeAnalyzer:
        def close(self):
            closed.append("closed")

    monkeypatch.setattr(server_module, "build_analyzer", lambda: FakeAnalyzer())

    with TestClient(create_app(auto_run_jobs=False)) as client:
        client.get("/api/health")
        server_module.get_or_build_analyzer(client.app)

    assert closed == ["closed"]
